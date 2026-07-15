"""Excel export for a computed cost estimate.

Workbook layout ported from the Self-Storage Analysis Suite's
"Quick Estimate" export (Real Estate Project repo, git commit cb54ad7).
"""

from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from estimator import QUALITY_MULT


def build_workbook(est: dict) -> Workbook:
    wb = Workbook()

    # ── Tab 1: Cost Estimate ──
    ws = wb.active
    ws.title = "Cost Estimate"

    header_font = Font(bold=True, size=14)
    bold_font = Font(bold=True)
    currency_fmt = '#,##0'
    currency_psf_fmt = '$#,##0.00'
    header_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    total_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

    ws.merge_cells("A1:C1")
    ws["A1"] = "Construction Cost Estimate — Quick Estimate"
    ws["A1"].font = header_font

    ws["A2"] = "Property Type:"
    ws["B2"] = est["building_label"]
    ws["A3"] = "Total SF:"
    ws["B3"] = est["sf"]
    ws["B3"].number_format = '#,##0'
    ws["A4"] = "City:"
    loc = est["location_factor"]
    matched = est["matched_city"]
    ws["B4"] = f"{est['city']} (Location Factor: {loc:.2f}x{' — ' + matched if matched else ''})"
    ws["A5"] = "Quality:"
    ws["B5"] = est["quality"]
    ws["A6"] = "Date:"
    ws["B6"] = date.today().strftime("%B %d, %Y")
    ws["A7"] = "Loan Rate:"
    ws["B7"] = f"{est['loan_rate']:.1f}% annual × {est['const_months']} months"

    row = 9
    for col, label in [(1, "Component"), (2, "$/SF"), (3, "Total Cost")]:
        cell = ws.cell(row=row, column=col, value=label)
        cell.font = bold_font
        cell.fill = header_fill
    row += 1

    for item in est["hard_items"]:
        ws.cell(row=row, column=1, value=item["name"])
        if item["psf_raw"] is not None:
            ws.cell(row=row, column=2, value=item["psf_raw"]).number_format = currency_psf_fmt
        else:
            ws.cell(row=row, column=2, value=item["note"])
        ws.cell(row=row, column=3, value=item["cost_raw"]).number_format = currency_fmt
        row += 1

    total_hard = est["total_hard"]
    sf = est["sf"]
    ws.cell(row=row, column=1, value="HARD COST SUBTOTAL").font = bold_font
    ws.cell(row=row, column=2, value=total_hard / sf if sf > 0 else 0).number_format = currency_psf_fmt
    ws.cell(row=row, column=2).font = bold_font
    ws.cell(row=row, column=3, value=total_hard).number_format = currency_fmt
    ws.cell(row=row, column=3).font = bold_font
    row += 2

    ws.cell(row=row, column=1, value="Soft Costs").font = bold_font
    ws.cell(row=row, column=2, value="% of Hard").font = bold_font
    row += 1

    for item in est["soft_items"]:
        ws.cell(row=row, column=1, value=item["name"])
        ws.cell(row=row, column=2, value=f"{item['pct_raw']:.1%}")
        ws.cell(row=row, column=3, value=item["cost_raw"]).number_format = currency_fmt
        row += 1

    total_soft = est["total_soft"]
    ws.cell(row=row, column=1, value="SOFT COST SUBTOTAL").font = bold_font
    ws.cell(row=row, column=3, value=total_soft).number_format = currency_fmt
    ws.cell(row=row, column=3).font = bold_font
    row += 2

    total_land = est["total_land"]
    if est["land_cost_input"] > 0:
        land_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
        for col, label in [(1, "Land & Acquisition Costs"), (2, "Note"), (3, "Total Cost")]:
            cell = ws.cell(row=row, column=col, value=label)
            cell.font = bold_font
            cell.fill = land_fill
        row += 1

        for item in est["land_items"]:
            ws.cell(row=row, column=1, value=item["name"])
            ws.cell(row=row, column=2, value=item["note"])
            ws.cell(row=row, column=3, value=item["cost_raw"]).number_format = currency_fmt
            row += 1

        ws.cell(row=row, column=1, value="LAND & ACQUISITION SUBTOTAL").font = bold_font
        ws.cell(row=row, column=3, value=total_land).number_format = currency_fmt
        ws.cell(row=row, column=3).font = bold_font
        row += 2

    grand_total = est["grand_total"]
    for col in range(1, 4):
        ws.cell(row=row, column=col).fill = total_fill
        ws.cell(row=row, column=col).font = bold_font
    ws.cell(row=row, column=1, value="TOTAL ESTIMATED COST")
    ws.cell(row=row, column=2, value=grand_total / sf if sf > 0 else 0).number_format = currency_psf_fmt
    ws.cell(row=row, column=3, value=grand_total).number_format = currency_fmt

    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18

    # ── Tab 2: Sources & Assumptions ──
    ws2 = wb.create_sheet("Sources & Assumptions")
    ws2["A1"] = "Sources & Assumptions"
    ws2["A1"].font = header_font

    ws2["A3"] = "Category"
    ws2["B3"] = "Assumption"
    ws2["C3"] = "Source"
    for col in range(1, 4):
        ws2.cell(row=3, column=col).font = bold_font
        ws2.cell(row=3, column=col).fill = header_fill

    quality = est["quality"]
    q_mult = QUALITY_MULT.get(quality, 1.0)

    sources = [
        ("Base Cost Data", "Per-SF costs are national averages for this property type",
         "RSMeans Building Construction Cost Data 2024/2025 — https://www.rsmeans.com"),
        ("Location Factor", f"{loc:.2f}x applied to {matched or 'national average'}",
         "USACE Area Cost Factors (PAX) — https://www.usace.army.mil/Cost-Engineering/"),
        ("Quality Multiplier", f"{quality} = {q_mult:.2f}x adjustment",
         "RSMeans quality adjustment methodology — Economy (0.85x), Average (1.00x), Premium (1.15x)"),
        ("A&E (5%)", "Architectural & Engineering fees as % of hard costs",
         "AIA Compensation Report — https://www.aia.org/resources/compensation-report"),
        ("Permits & Impact Fees (2.5%)", "Building permits and municipal impact fees",
         "National Association of Home Builders — https://www.nahb.org/advocacy/impact-fees"),
        ("Geotech / Environmental (0.8%)", "Geotechnical study, Phase I/II ESA",
         "ASTM E1527-21 standard practice — typical Phase I ESA cost ranges"),
        ("Survey & Land Planning (0.4%)", "ALTA survey, site planning",
         "NSPS/ALTA Standards — https://www.nsps.us.com"),
        ("Legal & Closing (0.8%)", "Legal fees, title, closing costs",
         "Industry standard — varies by deal complexity and jurisdiction"),
        ("Builder's Risk Insurance (0.7%)", "Construction-period insurance coverage",
         "IRMI (International Risk Management Institute) — https://www.irmi.com"),
        ("Construction Loan Interest", f"{est['loan_rate']:.1f}% annual rate × {est['const_months']} months × 50% avg draw",
         "User-entered rate. Formula: hard costs × rate × (months/12) × 0.5 average outstanding balance"),
        ("Property Tax During Const. (0.8%)", "Ad valorem tax on land during construction",
         "County assessor rates — varies by jurisdiction"),
        ("Contingency (7.5%)", "Unforeseen conditions, change orders",
         "AACE International Recommended Practice 18R-97 — Class 3 estimate contingency range"),
    ]

    if est["land_cost_input"] > 0:
        sources += [
            ("Land Purchase Price", "User-entered value",
             "User input — verify with purchase agreement or broker opinion of value"),
            ("Title & Closing (~1.5%)", "Title insurance, recording fees, escrow, legal review",
             "ALTA best practices — typical range 1-2% of land value"),
            ("Phase I ESA (~$4,000)", "Environmental site assessment per ASTM E1527-21",
             "ASTM E1527-21 — https://www.astm.org/e1527-21.html — typical range $3,000-$5,000"),
            ("ALTA Survey (~$7,500)", "Boundary and topographic survey",
             "NSPS/ALTA Standards — https://www.nsps.us.com — typical range $5,000-$10,000"),
            ("Utility Tap Fees (~$15,000)", "Water, sewer, electric connection fees",
             "Municipal fee schedules — varies widely by jurisdiction ($8,000-$25,000 typical)"),
        ]

    for i, (cat, assumption, source) in enumerate(sources, start=4):
        ws2.cell(row=i, column=1, value=cat)
        ws2.cell(row=i, column=2, value=assumption)
        ws2.cell(row=i, column=3, value=source)

    disclaimer_row = len(sources) + 6
    ws2.cell(row=disclaimer_row, column=1, value="DISCLAIMER").font = bold_font
    ws2.cell(row=disclaimer_row + 1, column=1,
             value="This is a preliminary estimate based on published cost data and standard industry assumptions. "
                   "Actual costs will vary based on site conditions, local labor markets, material pricing, "
                   "and project-specific requirements. Verify all figures with local contractors and consultants.")

    ws2.column_dimensions["A"].width = 35
    ws2.column_dimensions["B"].width = 55
    ws2.column_dimensions["C"].width = 75

    return wb


def export_filename(city: str) -> str:
    safe_city = (city or "national").replace(" ", "_").replace(",", "")
    return f"quick_estimate_{safe_city}_{date.today().strftime('%b-%d-%y')}.xlsx"
